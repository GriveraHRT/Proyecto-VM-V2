import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io, os

st.set_page_config(
    page_title="Estadística de Laboratorio",
    page_icon="🔬",
    layout="centered"
)

st.title("🔬 Procesador de Estadística de Laboratorio")
st.caption(
    "Sube el archivo de datos crudo (tal como lo entrega el sistema). "
    "La app depura los datos y entrega un Excel con 3 hojas: "
    "**Gral** (datos ordenados), **Centinela** y **Urgencia** (conteos epidemiológicos)."
)
st.divider()


# ╔══════════════════════════════════════════════════════════╗
# ║  CONSTANTES                                              ║
# ╚══════════════════════════════════════════════════════════╝

VIRUS_MAP = {
    'VIRUS RESPIRATORIO SINCICIAL': 'VRS',
    'ADENOVIRUS':                   'Adenovirus',
    'PARAINFLUENZA':                'Parainfluenza',
    'INFLUENZA A':                  'Influenza A',
    'INFLUENZA B':                  'Influenza B',
    'METAPNEUMOVIRUS':              'Metapneumovirus',
    'RHINOVIRUS':                   'Rhinovirus',
    'SARS COV 2 (COVID-19)':        'Covid-19',
}
VIRUS_ORDER = list(VIRUS_MAP.values())

ALL_AGES = ['Total', 'Menor 1 año', '1-4 años', '5-14 años',
            '15-54 años', '55-64 años', '65 y más']

CENTINELA_PROC = ['CESFAM CERRO ALTO']
URGENCIA_PROC  = ['URGENCIA PEDIATRIA', 'URGENCIA ADULTO', 'INGRESO MATER']

# Filas de la plantilla (1-indexed) para cada virus
VIRUS_ROWS = {
    'VRS': 12, 'Adenovirus': 13, 'Parainfluenza': 14, 'Influenza A': 15,
    'Influenza B': 16, 'Metapneumovirus': 17, 'Rhinovirus': 18,
    'Covid-19': 19, 'Negativos': 20, '_Total': 21,
}
TABLE2_OFFSET = 24  # la tabla de coinfecciones empieza 24 filas más abajo

# Columnas (M, F) para cada rango etario
AGE_COLS = {
    'Total':       (2, 3),
    'Menor 1 año': (4, 5),
    '1-4 años':    (6, 7),
    '5-14 años':   (8, 9),
    '15-54 años': (10, 11),
    '55-64 años': (12, 13),
    '65 y más':   (14, 15),
}


# ╔══════════════════════════════════════════════════════════╗
# ║  UTILIDADES                                              ║
# ╚══════════════════════════════════════════════════════════╝

def rango_etario(edad):
    try:
        e = float(edad)
        if e < 1:   return "Menor 1 año"
        if e <= 4:  return "1-4 años"
        if e <= 14: return "5-14 años"
        if e <= 54: return "15-54 años"
        if e <= 64: return "55-64 años"
        return "65 y más"
    except Exception:
        return "Sin dato"


def nueva_tabla():
    return {v: {a: {'M': 0, 'F': 0} for a in ALL_AGES}
            for v in VIRUS_ORDER + ['Negativos']}


def sumar(tabla, virus, edad, sexo):
    if virus in tabla:
        tabla[virus]['Total'][sexo] += 1
        if edad in tabla[virus]:
            tabla[virus][edad][sexo] += 1


# ╔══════════════════════════════════════════════════════════╗
# ║  PROCESAMIENTO PRINCIPAL                                 ║
# ╚══════════════════════════════════════════════════════════╝

def procesar(df):
    col_map = {c.strip().lower(): c for c in df.columns}

    def col(*nombres):
        for n in nombres:
            if n in col_map:
                return col_map[n]
        return None

    col_orden = col("orden", "nº orden", "numero orden", "n orden", "n° orden")
    col_valor = col("valor")
    col_edad  = col("edad")
    col_sexo  = col("sexo")
    col_proc  = col("procedencia")
    col_prest = col("prestación estructura", "prestacion estructura")

    faltantes = [n for n, c in [
        ("Orden", col_orden), ("Valor", col_valor), ("Edad", col_edad),
        ("Sexo", col_sexo), ("Procedencia", col_proc),
        ("Prestación Estructura", col_prest),
    ] if not c]
    if faltantes:
        return None, f"No se encontraron las columnas: {', '.join(faltantes)}"

    n_orig = len(df)

    # Eliminar filas de tipo "MUESTRA" (Hisopado/Aspirado nasofaríngeo, etc.)
    mask_muestra = df[col_prest].astype(str).str.strip().str.upper() == 'MUESTRA'
    df = df[~mask_muestra].copy()
    n_muestra = int(mask_muestra.sum())

    df['Rango Etario'] = df[col_edad].apply(rango_etario)

    def es_positivo(v):
        return str(v).strip().upper() == 'POSITIVO'

    cent_set = {p.upper() for p in CENTINELA_PROC}
    urg_set  = {p.upper() for p in URGENCIA_PROC}

    gral_rows, flags = [], []
    t1c, t2c = nueva_tabla(), nueva_tabla()  # Centinela: únicas / coinfecciones
    t1u, t2u = nueva_tabla(), nueva_tabla()  # Urgencia:  únicas / coinfecciones

    for _, grupo in df.groupby(col_orden, sort=False):
        positivos = grupo[grupo[col_valor].apply(es_positivo)]

        sexo = 'M' if 'MASCUL' in str(grupo.iloc[0][col_sexo]).upper() else 'F'
        edad = grupo.iloc[0]['Rango Etario']
        proc = str(grupo.iloc[0][col_proc]).strip().upper()

        es_centinela = proc in cent_set
        es_urgencia  = proc in urg_set

        if len(positivos) == 0:
            gral_rows.append(grupo.iloc[0])
            flags.append(False)
            if es_centinela: sumar(t1c, 'Negativos', edad, sexo)
            if es_urgencia:  sumar(t1u, 'Negativos', edad, sexo)

        elif len(positivos) == 1:
            fila = positivos.iloc[0]
            gral_rows.append(fila)
            flags.append(False)
            virus = VIRUS_MAP.get(str(fila[col_prest]).strip().upper())
            if virus:
                if es_centinela: sumar(t1c, virus, edad, sexo)
                if es_urgencia:  sumar(t1u, virus, edad, sexo)

        else:
            for _, fila in positivos.iterrows():
                gral_rows.append(fila)
                flags.append(True)
                virus = VIRUS_MAP.get(str(fila[col_prest]).strip().upper())
                if virus:
                    if es_centinela: sumar(t2c, virus, edad, sexo)
                    if es_urgencia:  sumar(t2u, virus, edad, sexo)

    gral = pd.DataFrame(gral_rows).reset_index(drop=True)

    return {
        'gral': gral, 'flags': flags,
        't1c': t1c, 't2c': t2c, 't1u': t1u, 't2u': t2u,
        'n_orig': n_orig, 'n_muestra': n_muestra,
    }, None


# ╔══════════════════════════════════════════════════════════╗
# ║  GENERACIÓN DEL EXCEL                                    ║
# ╚══════════════════════════════════════════════════════════╝

def totales_tabla(tabla):
    tots = {a: {'M': 0, 'F': 0} for a in ALL_AGES}
    for v in VIRUS_ORDER + ['Negativos']:
        for a in ALL_AGES:
            tots[a]['M'] += tabla.get(v, {}).get(a, {}).get('M', 0)
            tots[a]['F'] += tabla.get(v, {}).get(a, {}).get('F', 0)
    return tots


def llenar_hoja_epi(ws, t1, t2):
    def sv(r, c, val):
        ws.cell(row=r, column=c).value = val or None

    for v in VIRUS_ORDER + ['Negativos']:
        r1 = VIRUS_ROWS[v]
        r2 = r1 + TABLE2_OFFSET
        for edad, (cm, cf) in AGE_COLS.items():
            d1 = t1.get(v, {}).get(edad, {'M': 0, 'F': 0})
            sv(r1, cm, d1['M']); sv(r1, cf, d1['F'])
            d2 = t2.get(v, {}).get(edad, {'M': 0, 'F': 0})
            sv(r2, cm, d2['M']); sv(r2, cf, d2['F'])

    tr1, tr2 = VIRUS_ROWS['_Total'], VIRUS_ROWS['_Total'] + TABLE2_OFFSET
    tot1, tot2 = totales_tabla(t1), totales_tabla(t2)
    for edad, (cm, cf) in AGE_COLS.items():
        sv(tr1, cm, tot1[edad]['M']); sv(tr1, cf, tot1[edad]['F'])
        sv(tr2, cm, tot2[edad]['M']); sv(tr2, cf, tot2[edad]['F'])


@st.cache_resource
def cargar_plantilla():
    ruta = os.path.join(os.path.dirname(__file__), 'plantilla_informe.xlsx')
    with open(ruta, 'rb') as f:
        return f.read()


def construir_excel(res):
    wb = load_workbook(io.BytesIO(cargar_plantilla()))

    # ── Hoja Gral ──
    ws = wb.create_sheet('Gral', 0)
    gral, flags = res['gral'], res['flags']
    cols = list(gral.columns)
    ws.append(cols)

    h_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    h_font = Font(color="FFFFFF", bold=True, name="Arial", size=10)
    r_fill = PatternFill("solid", start_color="FFCCCC", end_color="FFCCCC")
    r_font = Font(color="CC0000", bold=True, name="Arial", size=10)
    normal = Font(name="Arial", size=10)
    borde  = Border(**{s: Side(style="thin", color="CCCCCC")
                       for s in ("left", "right", "top", "bottom")})

    for c in ws[1]:
        c.fill = h_fill; c.font = h_font; c.border = borde
        c.alignment = Alignment(horizontal="center")

    for _, fila in gral.iterrows():
        valores = []
        for v in fila:
            if isinstance(v, pd.Timestamp):
                v = v.to_pydatetime()
            valores.append(v if pd.notna(v) else None)
        ws.append(valores)

    for i, multi in enumerate(flags, start=2):
        for ci in range(1, len(cols) + 1):
            c = ws.cell(i, ci)
            c.border = borde
            c.font   = r_font if multi else normal
            if multi:
                c.fill = r_fill

    for ci in range(1, len(cols) + 1):
        letra = get_column_letter(ci)
        ancho = max(len(str(ws.cell(r, ci).value or ""))
                    for r in range(1, min(ws.max_row + 1, 50)))
        ws.column_dimensions[letra].width = min(max(ancho + 4, 12), 40)

    ws.auto_filter.ref = ws.dimensions

    # ── Hojas Centinela / Urgencia ──
    llenar_hoja_epi(wb['Centinela'], res['t1c'], res['t2c'])
    llenar_hoja_epi(wb['Urgencia'],  res['t1u'], res['t2u'])

    salida = io.BytesIO()
    wb.save(salida)
    salida.seek(0)
    return salida


# ╔══════════════════════════════════════════════════════════╗
# ║  VISTA PREVIA                                            ║
# ╚══════════════════════════════════════════════════════════╝

def vista_previa(tabla, titulo):
    filas = [{'Virus': v,
              'M': tabla.get(v, {}).get('Total', {}).get('M', 0),
              'F': tabla.get(v, {}).get('Total', {}).get('F', 0)}
             for v in VIRUS_ORDER + ['Negativos']]
    df_p = pd.DataFrame(filas)
    df_p['Total'] = df_p['M'] + df_p['F']
    df_p = df_p[df_p['Total'] > 0]

    st.write(f"**{titulo}**")
    if df_p.empty:
        st.write("_(sin casos)_")
    else:
        st.dataframe(df_p, use_container_width=True, hide_index=True)


# ╔══════════════════════════════════════════════════════════╗
# ║  INTERFAZ                                                ║
# ╚══════════════════════════════════════════════════════════╝

archivo = st.file_uploader(
    "Selecciona el archivo Excel de datos crudos",
    type=["xls", "xlsx"],
    help="El archivo tal como lo exporta el sistema, sin procesar."
)

if archivo:
    with st.spinner("Leyendo archivo..."):
        try:
            df_raw = pd.read_excel(archivo)
        except Exception as e:
            st.error(f"No se pudo leer el archivo: {e}")
            st.stop()

    with st.spinner("Procesando datos..."):
        res, error = procesar(df_raw)
        if error:
            st.error(error)
            st.stop()

    n_multi = sum(res['flags'])

    st.divider()
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Filas originales",          res['n_orig'])
    c2.metric("Filas de muestra eliminadas", res['n_muestra'])
    c3.metric("Filas resultado (Gral)",    len(res['gral']))
    c4.metric("⚠️ Coinfecciones",           n_multi)

    if n_multi:
        st.warning(
            f"Se encontraron **{n_multi} filas** con coinfección "
            "(más de un virus positivo por orden). Quedan marcadas en rojo "
            "en la hoja **Gral**.",
            icon="⚠️"
        )

    st.subheader("Vista previa — Gral")
    st.dataframe(res['gral'].head(10), use_container_width=True, hide_index=True)

    st.divider()
    col_a, col_b = st.columns(2)
    with col_a:
        st.subheader("🏥 Centinela")
        st.caption("CESFAM Cerro Alto")
        vista_previa(res['t1c'], "Infecciones únicas")
        vista_previa(res['t2c'], "Coinfecciones")
    with col_b:
        st.subheader("🚨 Urgencia")
        st.caption("Urg. Pediatría · Urg. Adulto · Ingreso Mater")
        vista_previa(res['t1u'], "Infecciones únicas")
        vista_previa(res['t2u'], "Coinfecciones")

    with st.spinner("Generando Excel..."):
        excel_bytes = construir_excel(res)

    st.divider()
    st.download_button(
        label="⬇️ Descargar Excel (Gral + Centinela + Urgencia)",
        data=excel_bytes,
        file_name=archivo.name.rsplit(".", 1)[0] + "_procesado.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary"
    )
